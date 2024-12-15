###----------------------------------------------------------------
### Utility Functions
###---------------------------------------------------------

import os
import numpy as np
import cv2
from fuzzywuzzy import process, fuzz
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import io


# Replace numbers with letters
# Define the mappings from numbers to similar-looking letters
number_to_letter_map = {
    '0': 'O',  # zero to uppercase 'O'
    '1': 'I',  # one to uppercase 'I'
    '2': 'Z',  # two to uppercase 'Z'
    '3': 'E',  # three to uppercase 'E'
    '4': 'A',  # four to uppercase 'A'
    '5': 'S',  # five to uppercase 'S'
    '6': 'G',  # six to uppercase 'G'
    '7': 'T',  # seven to uppercase 'T'
    '8': 'B',  # eight to uppercase 'B'
    '9': 'P',  # nine to uppercase 'P'
}


# Replace numbers with similar-looking letters
def replace_numbers_with_letters(text):
    # Substitute each number with its corresponding letter
    for num, letter in number_to_letter_map.items():
        text = text.replace(num, letter)
    return text


# Reward for same initial letter
def custom_fuzzy_match(query, bird_names, prefix_weight=1.25):
    best_match = None
    highest_score = 0

    for bird_name in bird_names:
        # Berechnen Sie die Fuzzy-Übereinstimmung
        score = fuzz.ratio(query, bird_name)

        # Fügen Sie eine Gewichtung hinzu, wenn der Anfangsbuchstabe übereinstimmt
        if bird_name.lower().startswith(query[0].lower()):
            score *= prefix_weight

        # Aktualisieren Sie das beste Ergebnis
        if score > highest_score:
            highest_score = score
            best_match = bird_name

    return best_match, highest_score


# Function to determine the predominant color in an image
def check_text_color(image, blue_ratio_threshold=0.02):
    # Convert to HSV
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Get the height, width, and channels of the image
    height, width, _ = image.shape

    # Define the HSV range for blue
    lower_blue = np.array([100, 50, 50])  # Lower bound of the blue range
    upper_blue = np.array([140, 255, 255])

    # Create mask
    mask_blue = cv2.inRange(hsv, lower_blue, upper_blue)

    # Count the number of pixels in each mask
    blue_count = np.sum(mask_blue > 0)

    # Determine if text is predominantly black or blue
    total_pixels = height * width
    blue_ratio = blue_count / total_pixels

    # Only Blue Pixels
    imask = mask_blue > 0
    blue = np.full((height, width, _), 255, dtype=np.uint8)
    blue[imask] = image[imask]

    # plt.imshow(cv.cvtColor(blue, cv.COLOR_BGR2RGB))
    # plt.show()

    # Print the results for insight
    return {
        'blue_ratio': blue_ratio,
        'predominant_color': 'blue' if blue_ratio > blue_ratio_threshold else 'black'
    }


# Function to replace all black pixels with white in an image
def remove_black_pixels(image):  # detect all non-black pixels

    # Get the height, width, and channels of the image
    height, width, _ = image.shape

    # Define the HSV range for all colors except black and near-black
    lower_black = np.array([0, 50, 50])  # Lower bound to exclude black and near-black
    upper_black = np.array([179, 255, 255])  # Upper bound for all colors

    ## Convert to HSV
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, lower_black, upper_black)

    ## Slice
    imask = mask > 0
    white = np.full((height, width, _), 255, dtype=np.uint8)
    white[imask] = image[imask]

    return white


def detectQuotationMarks(image, relative_non_white_threshold=0.1157):
    # Convert the image to HSV color space
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Define the lower and upper bounds for "nearly white"
    lower_white = np.array([0, 0, 200])  # Low saturation and high value
    upper_white = np.array([180, 50, 255])  # Wide range of hue and low saturation

    # Create a mask for pixels that are considered white or nearly white
    mask = cv2.inRange(hsv, lower_white, upper_white)

    # Invert the mask to count non-white pixels
    non_white_pixels = np.sum(mask == 0)

    # Get the total number of pixels in the image
    total_pixels = hsv.shape[0] * hsv.shape[1]

    # Calculate the relative amount of non-white pixels
    relative_non_white = non_white_pixels / total_pixels

    # Determine if the image contains text or just quotation marks based on the ratio
    if relative_non_white < relative_non_white_threshold:
        return {
            'Quotationmark': True,
            'relativeNonWhite': relative_non_white
        }
    else:
        return {
            'Quotationmark': False,
            'relativeNonWhite': relative_non_white
        }


def initialize_excel(num_pages, start_number, output_path):
    """
    Initializes an Excel workbook with the specified number of sheets, each with numbers starting from the given number in the first column.

    Parameters:
        num_pages (int): Number of sheets (pages) to create in the workbook.
        start_number (int): The starting number for the first column.
        output_path (str): File path to save the initialized workbook.

    Returns:
        None
    """
    # Create a new workbook
    workbook = Workbook()

    # Create sheets with the specified format
    for page in range(num_pages):
        if page == 0:
            sheet = workbook.active
            sheet.title = f"{page + 1}.png"
        else:
            sheet = workbook.create_sheet(title=f"{page + 1}.png")

        # Fill the first column with numbers starting from the specified start_number
        for row in range(1, 51):
            sheet.cell(row=row, column=1, value=start_number + (page * 50) + (row - 1))

    # Save the workbook to the specified path
    workbook.save(output_path)
    print(f"Excel workbook initialized and saved to {output_path}")


def write_ocr_outputs_to_excel(ocr_outputs, excel_path):
    """
    Writes OCR outputs to an existing Excel workbook.

    Parameters:
        ocr_outputs (list): List of dictionaries containing OCR data.
        excel_path (str): Path to the Excel workbook to update.

    Returns:
        None
    """
    # Load the existing workbook
    workbook = openpyxl.load_workbook(excel_path)

    # Iterate over OCR outputs and write data to the appropriate sheet
    for entry in ocr_outputs:
        page_number = entry.get("page_number", 1)  # Default to page 1 if not specified
        line_number = entry.get("line_number", 1)  # Default to line 1 if not specified

        sheet_name = f"{page_number}.png"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Adjust row height
            sheet.row_dimensions[line_number].height = 40.0  # Adjust height to 30 points (around 40 pixels)

            # Adjust column width to 15 for each column from 1 to 5 (A-D)
            for col in range(1, 5):  # Columns A to E
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            sheet.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 55


            # Write detected bird name, score, and true text to the appropriate columns
            sheet.cell(row=line_number, column=2, value=entry.get("best_match_customFuzzy"))
            sheet.cell(row=line_number, column=3, value=entry.get("score_customFuzzy"))
            sheet.cell(row=line_number, column=4, value=entry.get("true_text"))

            # Add image to column 5
            image = entry.get("image")
            if image is not None:
                try:
                    # Convert the PIL Image to a file-like object (BytesIO)
                    img_byte_array = io.BytesIO()
                    image.save(img_byte_array, format='PNG')  # Save the image in memory as PNG
                    img_byte_array.seek(0)  # Seek to the beginning of the byte array

                    # Create an ExcelImage object from the file-like object
                    excel_image = ExcelImage(img_byte_array)

                    # Specify the position (column 5, row = line_number)
                    cell_address = f"E{line_number}"  # E is the 5th column
                    excel_image.anchor = cell_address

                    # Add the image to the sheet
                    sheet.add_image(excel_image)
                except Exception as e:
                    print(f"Error adding image for page {page_number}, line {line_number}: {e}")

    # Save changes to the workbook
    try:
        workbook.save(excel_path)
        print(f"OCR outputs written to {excel_path}")
    except Exception as e:
        print(f"Error saving workbook: {e}")

def add_color_to_scores(score_column, threshold, excel_path):
    """
    Adds color to cells in the specified score column based on a threshold.

    Parameters:
        excel_path (str): Path to the Excel workbook.
        score_column (int): The column number containing the scores (e.g., 3 for column 'C').
        threshold (float): The score threshold for coloring.

    Returns:
        None
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_path)

    # Define red and green fill styles
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Iterate over all sheets
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate over all rows in the sheet
        for row in range(1, sheet.max_row + 1):
            score_cell = sheet.cell(row=row, column=score_column)

            # Check if the cell contains a numeric score
            if score_cell.value is not None and isinstance(score_cell.value, (int, float)):
                if score_cell.value < threshold:
                    score_cell.fill = red_fill
                else:
                    score_cell.fill = green_fill

    # Save the workbook
    workbook.save(excel_path)
    print(f"Colors added to scores in {excel_path}")